from constans import *
import openpyxl

user_dict = {}



res = list()
user_dict = {}


class User:
    def __init__(self, name):
        self.name = name
        self.password = "password"
        self.file = "File"



# Handle '/start' and '/help'
@bot.message_handler(commands=['help', 'start'])
def send_welcome(message):
    msg = bot.reply_to(message, """\
Привет, это бот, который присылает негатив. Введите свою кодовую фразу.
""")
    bot.register_next_step_handler(msg, process_name_step)

def process_name_step(message):
    chat_id = message.chat.id
    password = message.text
    user = User(password)
    user_dict[chat_id] = user
    user.password = password
    
    name_user=message.from_user.first_name
    user_id=message.from_user.id
    try:
        if user.password == "pishnyij":
            bot.send_message(chat_id , 'Здравствуйте, выберите, что вас интересует', reply_markup=markup_for_menu)
            file = "./pishnyij.xlsx"
            bot.send_message(308984767, f'name: {name_user} зашел chat_id{chat_id} user_id{user_id}')
            user.file = file
            user = user_dict[chat_id]
            print('User name client : ' , name_user)
            print('User chat id : ' , chat_id)
            print('User user_id : ' , user_id)
            
        if user.password == "churilo":
            bot.send_message(chat_id , 'Здравствуйте, выберите, что вас интересует', reply_markup=markup_for_menu)
            bot.send_message(308984767, f'name: {name_user} зашел chat_id{chat_id} user_id{user_id}')
            file = "./churilo.xlsx"
            user.file = file
            user = user_dict[chat_id]
            
            print('User name client : ' , name_user)
            print('User chat id : ' , chat_id)
            print('User user_id : ' , user_id)
            
            
        
        if user.password == "duplij":
            
            print('User name client : ' , name_user)
            print('User chat id : ' , chat_id)
            print('User user_id : ' , user_id)
            bot.send_message(chat_id , 'Здравствуйте, выберите, что вас интересует', reply_markup=markup_for_menu)
            bot.send_message(308984767, f'name: {name_user} зашел chat_id{chat_id} user_id{user_id}')
            file = "./duplij.xlsx"
            user.file = file
            user = user_dict[chat_id]
            
    except Exception as e:
        bot.reply_to(message, 'oooops')

def menu(message):
    chat_id = message.chat.id
    bot.send_message(chat_id , 'Здравствуйте, выберите, что вас интересует', reply_markup=markup_for_menu)



@bot.message_handler(regexp="Facebook")
def facebook(message):
    chat_id = message.chat.id
    user = user_dict[chat_id] 
    print("File", user.file)
    wb = openpyxl.load_workbook(user.file)
    ws = wb.active
    for i in range(2, len(ws['A'])+1):
    
        site = ws[f'A{i}'].value
        bot.send_message(message.chat.id, site)
            
@bot.message_handler(regexp="Youtube")
def youtube(message):
    chat_id = message.chat.id
    user = user_dict[chat_id] 
    print("File", user.file)
    wb = openpyxl.load_workbook(user.file)
    ws = wb.active
    for i in range(2, len(ws['B'])+1):
    
        site = ws[f'B{i}'].value
        bot.send_message(message.chat.id, site)

@bot.message_handler(regexp="На других сервисах")
def na_drugin(message):
    chat_id = message.chat.id
    user = user_dict[chat_id] 
    
    print("File", user.file)
    wb = openpyxl.load_workbook(user.file)
    ws = wb.active
    for i in range(2, len(ws['C'])+1):
    
        site = ws[f'C{i}'].value
        bot.send_message(message.chat.id, site)
    

@bot.message_handler(regexp="Вчера")
def davnije(message):
    chat_id = message.chat.id
    user = user_dict[chat_id] 
    print("File", user.file)
    wb = openpyxl.load_workbook(user.file)
    ws = wb.active
    for i in range(2, len(ws['D'])+1):
    
        site = ws[f'D{i}'].value
        bot.send_message(message.chat.id, site)




@bot.message_handler(regexp="Сегодня")
def vchera(message):
    chat_id = message.chat.id
    user = user_dict[chat_id] 
    print("File", user.file)
    wb = openpyxl.load_workbook(user.file)
    ws = wb.active
    for i in range(2, len(ws['E'])+1):
    
        site = ws[f'E{i}'].value
        bot.send_message(message.chat.id, site)

@bot.message_handler(regexp="Все")
def vse(message):
    chat_id = message.chat.id
    user = user_dict[chat_id] 
    print("File", user.file)
    wb = openpyxl.load_workbook(user.file)
    ws = wb.active
    for i in range(2, len(ws['F'])+1):
    
        site = ws[f'F{i}'].value
        bot.send_message(message.chat.id, site)
if __name__ == '__main__':

# Enable saving next step handlers to file "./.handlers-saves/step.save".
# Delay=2 means that after any change in next step handlers (e.g. calling register_next_step_handler())
# saving will hapen after delay 2 seconds.
    bot.enable_save_next_step_handlers(delay=6)

# Load next_step_handlers from save file (default "./.handlers-saves/step.save")
# WARNING It will work only if enable_save_next_step_handlers was called!
    bot.load_next_step_handlers()

    bot.polling(none_stop=True)